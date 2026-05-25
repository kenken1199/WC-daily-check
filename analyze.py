import platform
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.font_manager as fm
import numpy as np
from scipy import stats
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import datetime
from io import BytesIO
import os
import sys


# =========================
# ■ 定数
# =========================
MIN_OK_COUNT = 2    # 統計分析に必要な最小OKデータ数
MA_WINDOW    = 500  # 時系列グラフの移動平均ウィンドウ幅（デフォルト）


def _hinshoku_display(hinshoku_num):
    return f"品種番号{hinshoku_num}" if hinshoku_num is not None else None


def _setup_japanese_font():
    candidates = {
        "Windows": ["MS Gothic", "Yu Gothic", "Meiryo"],
        "Darwin":  ["Hiragino Sans", "Hiragino Maru Gothic Pro", "AppleGothic"],
        "Linux":   ["Noto Sans CJK JP", "IPAGothic", "IPAPGothic"],
    }.get(platform.system(), [])
    available = {f.name for f in fm.fontManager.ttflist}
    for font in candidates:
        if font in available:
            plt.rcParams["font.family"] = font
            return

_setup_japanese_font()

app_root = None
csv_file_to_process = None


# =========================
# ■ ロットプレビューダイアログ
# =========================
class LotPreviewDialog(tk.Toplevel):

    def __init__(self, parent, df, hinshoku_num=None):
        super().__init__(parent)
        self.title("ロット分割プレビュー")
        self.result = None
        self.df = df.copy()
        self.hinshoku_num = hinshoku_num
        self.resizable(True, True)
        self.grab_set()

        # --- しきい値入力 ---
        frame_top = ttk.Frame(self, padding=10)
        frame_top.pack(fill="x")

        ttk.Label(frame_top, text="分割しきい値（分）:").pack(side="left")
        self.threshold_var = tk.IntVar(value=30)
        ttk.Spinbox(
            frame_top, from_=1, to=480,
            textvariable=self.threshold_var, width=6
        ).pack(side="left", padx=5)
        ttk.Button(frame_top, text="更新", command=self._update_preview).pack(side="left", padx=5)

        self.lot_label_var = tk.StringVar()
        ttk.Label(frame_top, textvariable=self.lot_label_var, foreground="navy").pack(side="left", padx=15)

        # --- 移動平均ウィンドウ入力 ---
        frame_ma = ttk.Frame(self, padding=(10, 0, 10, 5))
        frame_ma.pack(fill="x")

        ttk.Label(frame_ma, text="移動平均ウィンドウ:").pack(side="left")
        self.ma_window_var = tk.IntVar(value=MA_WINDOW)
        ttk.Spinbox(
            frame_ma, from_=10, to=5000,
            textvariable=self.ma_window_var, width=7,
            command=self._update_ma_label
        ).pack(side="left", padx=5)
        ttk.Label(frame_ma, text="点").pack(side="left")
        self.ma_approx_var = tk.StringVar()
        ttk.Label(frame_ma, textvariable=self.ma_approx_var, foreground="gray").pack(side="left", padx=8)

        # --- Treeview ---
        frame_tree = ttk.Frame(self, padding=10)
        frame_tree.pack(fill="both", expand=True)

        cols = ("品種", "ロット", "開始時刻", "終了時刻", "総件数", "OKデータ件数", "状態")
        self.tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=10)
        col_widths = {"品種": 140, "ロット": 70, "開始時刻": 150, "終了時刻": 150,
                      "総件数": 80, "OKデータ件数": 110, "状態": 110}
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="center", width=col_widths[col])

        # 警告色タグ（OKデータ不足のロットを赤系で強調）
        self.tree.tag_configure("skip", background="#FFE4E1", foreground="#9C0006")

        scrollbar = ttk.Scrollbar(frame_tree, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # --- 注意書きラベル ---
        self.warn_label_var = tk.StringVar()
        warn_label = ttk.Label(self, textvariable=self.warn_label_var,
                               foreground="#9C0006", padding=(10, 0))
        warn_label.pack(fill="x")

        # --- ボタン ---
        frame_btn = ttk.Frame(self, padding=(10, 5, 10, 10))
        frame_btn.pack(fill="x")
        ttk.Button(frame_btn, text="この分割でOK", command=self._on_ok).pack(side="left", padx=5)
        ttk.Button(frame_btn, text="ロット分割しない", command=self._on_no_split).pack(side="left", padx=5)
        ttk.Button(frame_btn, text="手動分割する", command=self._on_manual).pack(side="left", padx=5)
        ttk.Button(frame_btn, text="キャンセル", command=self._on_cancel).pack(side="right", padx=5)

        self._update_preview()

        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

    def _compute_lots(self, threshold_min):
        df = self.df.sort_values("日付時刻").copy()
        df["時間差(分)"] = df["日付時刻"].diff().dt.total_seconds() / 60
        df["ロット"] = (df["時間差(分)"] > threshold_min).cumsum() + 1
        return df

    def _update_ma_label(self):
        try:
            ma_win = self.ma_window_var.get()
        except tk.TclError:
            return
        df_sorted = self.df.sort_values("日付時刻")
        valid_times = df_sorted["日付時刻"].dropna()
        if len(valid_times) >= 2:
            total_min = (valid_times.max() - valid_times.min()).total_seconds() / 60
            if total_min > 0:
                rate = len(df_sorted) / total_min
                approx_min = ma_win / rate
                self.ma_approx_var.set(f"≈ {approx_min:.0f}分ぶん（{ma_win}個ぶん）")
                return
        self.ma_approx_var.set(f"（{ma_win}個ぶん）")

    def _update_preview(self):
        threshold = self.threshold_var.get()
        df = self._compute_lots(threshold)

        for item in self.tree.get_children():
            self.tree.delete(item)

        lot_count = df["ロット"].nunique()
        skip_count = 0

        hinshoku_display = _hinshoku_display(self.hinshoku_num) or "-"

        for lot, group in df.groupby("ロット"):
            start = group["日付時刻"].min()
            end = group["日付時刻"].max()
            total = len(group)
            ok_count = int((group["ランクコード"] == "2").sum())

            if ok_count < MIN_OK_COUNT:
                state = "⚠ スキップ予定"
                tags = ("skip",)
                skip_count += 1
            else:
                state = "✓ 分析対象"
                tags = ()

            self.tree.insert("", "end", tags=tags, values=(
                hinshoku_display,
                f"ロット{lot}",
                start.strftime("%Y-%m-%d %H:%M") if pd.notna(start) else "-",
                end.strftime("%Y-%m-%d %H:%M") if pd.notna(end) else "-",
                total,
                ok_count,
                state,
            ))

        # 上部ラベル
        if skip_count > 0:
            self.lot_label_var.set(
                f"→ {lot_count} ロット検出（うち {skip_count} ロットはOKデータ不足でスキップ予定）"
            )
            self.warn_label_var.set(
                f"※ OKデータが {MIN_OK_COUNT} 件未満のロットは統計計算ができないため、"
                f"Excelファイルは作成されません。しきい値を変更してロットを統合することも検討してください。"
            )
        else:
            self.lot_label_var.set(f"→ {lot_count} ロット検出")
            self.warn_label_var.set("")

        self._df_with_lots = df
        self._update_ma_label()

    def _on_ok(self):
        self.ma_window = self.ma_window_var.get()
        self.result = ("ok", self._df_with_lots)
        self.destroy()

    def _on_no_split(self):
        self.ma_window = self.ma_window_var.get()
        df = self.df.copy()
        df["ロット"] = 1
        self.result = ("ok", df)
        self.destroy()

    def _on_manual(self):
        self.result = ("manual", None)
        self.destroy()

    def _on_cancel(self):
        self.result = ("cancel", None)
        self.destroy()



# =========================
# ■ 規格値入力ダイアログ
# =========================
class SpecInputDialog(tk.Toplevel):
    """
    result: "skip" | (nominal_or_None, ucl_or_None, lcl_or_None)
      nominal: 基準値 — None のとき先頭5個OK品平均を自動使用
      ucl/lcl: 上限値/下限値 — None のとき規格線なし（非対称も可）
    """

    def __init__(self, parent, hinshoku_num=None, lot=None, ok_count=None):
        super().__init__(parent)
        lot_label = f"ロット{lot}" if lot is not None else ""
        self.title(f"規格値入力　{lot_label}".strip())
        self.result = "skip"
        self.lot_label = None
        self.product_name = None
        self.resizable(False, False)
        self.grab_set()

        frame = ttk.Frame(self, padding=18)
        frame.pack(fill="both", expand=True)

        info_parts = []
        if hinshoku_num is not None:
            info_parts.append(f"品種番号: {hinshoku_num}")
        if lot is not None:
            info_parts.append(f"ロット {lot}")
        if ok_count is not None:
            info_parts.append(f"OKデータ {ok_count}件")
        if info_parts:
            ttk.Label(frame, text="  /  ".join(info_parts),
                      font=("", 10, "bold")).grid(row=0, column=0, columnspan=3,
                                                   sticky="w", pady=(0, 6))

        ttk.Label(frame, text="製品名 (任意):").grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.name_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.name_var, width=20).grid(
            row=1, column=1, columnspan=2, sticky="w", padx=5, pady=(4, 0))

        ttk.Label(frame, text="ロット名 (任意):").grid(row=2, column=0, sticky="w", pady=(8, 0))
        self.lot_label_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.lot_label_var, width=20).grid(
            row=2, column=1, sticky="w", padx=5, pady=(8, 0))
        ttk.Label(frame, text="※空白で「" + lot_label + "」を自動使用",
                  foreground="gray").grid(row=2, column=2, sticky="w", pady=(8, 0))

        ttk.Label(frame, text="基準値 (g):").grid(row=3, column=0, sticky="w", pady=(8, 0))
        self.nominal_var = tk.StringVar()
        self.nominal_var.trace_add("write", self._update_preview)
        ttk.Entry(frame, textvariable=self.nominal_var, width=12).grid(
            row=3, column=1, sticky="w", padx=5, pady=(8, 0))
        ttk.Label(frame, text="※空白で先頭5個の平均を自動使用",
                  foreground="gray").grid(row=3, column=2, sticky="w", pady=(8, 0))

        ttk.Label(frame, text="上側許容幅 (g):").grid(row=4, column=0, sticky="w", pady=(8, 0))
        self.upper_var = tk.StringVar()
        self.upper_var.trace_add("write", self._update_preview)
        ttk.Entry(frame, textvariable=self.upper_var, width=12).grid(
            row=4, column=1, sticky="w", padx=5, pady=(8, 0))
        ttk.Label(frame, text="基準値より何g上か（空白で規格線なし）",
                  foreground="gray").grid(row=4, column=2, sticky="w", pady=(8, 0))

        ttk.Label(frame, text="下側許容幅 (g):").grid(row=5, column=0, sticky="w", pady=(4, 0))
        self.lower_var = tk.StringVar()
        self.lower_var.trace_add("write", self._update_preview)
        ttk.Entry(frame, textvariable=self.lower_var, width=12).grid(
            row=5, column=1, sticky="w", padx=5, pady=(4, 0))
        ttk.Label(frame, text="基準値より何g下か",
                  foreground="gray").grid(row=5, column=2, sticky="w", pady=(4, 0))

        self.preview_var = tk.StringVar(value="→ UCL: −    LCL: −")
        ttk.Label(frame, textvariable=self.preview_var,
                  foreground="navy").grid(row=6, column=0, columnspan=3,
                                          pady=(10, 0), sticky="w")

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=7, column=0, columnspan=3, pady=(12, 0))
        ttk.Button(btn_frame, text="OK",     command=self._on_ok).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="スキップ", command=self._on_skip).pack(side="left", padx=5)

        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width()  - self.winfo_width())  // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

    def _update_preview(self, *_):
        try:
            upper = float(self.upper_var.get())
            lower = float(self.lower_var.get())
            nominal_str = self.nominal_var.get().strip()
            if nominal_str:
                n = float(nominal_str)
                self.preview_var.set(
                    f"→ UCL: {n + upper:.3f} g    LCL: {n - lower:.3f} g    幅: {upper + lower:.3f} g")
            else:
                self.preview_var.set(
                    f"→ UCL = 基準値 +{upper:.3f} g    LCL = 基準値 −{lower:.3f} g")
        except ValueError:
            self.preview_var.set("→ UCL: −    LCL: −")

    def _on_ok(self):
        nominal_str = self.nominal_var.get().strip()
        upper_str   = self.upper_var.get().strip()
        lower_str   = self.lower_var.get().strip()

        nominal = None
        if nominal_str:
            try:
                nominal = float(nominal_str)
            except ValueError:
                messagebox.showwarning("入力エラー", "基準値に数値を入力してください", parent=self)
                return

        upper = lower = None
        for val_str, label, attr in [(upper_str, "上側許容幅", "upper"),
                                     (lower_str, "下側許容幅", "lower")]:
            if val_str:
                try:
                    v = float(val_str)
                    if v <= 0:
                        messagebox.showwarning("入力エラー", f"{label}は正の値を入力してください", parent=self)
                        return
                    if attr == "upper":
                        upper = v
                    else:
                        lower = v
                except ValueError:
                    messagebox.showwarning("入力エラー", f"{label}に数値を入力してください", parent=self)
                    return

        self.result = (nominal, upper, lower)
        self.lot_label = self.lot_label_var.get().strip() or None
        self.product_name = self.name_var.get().strip() or None
        self.destroy()

    def _on_skip(self):
        self.result = "skip"
        self.destroy()


# =========================
# ■ CSV正規化
# =========================
_ISHIDA_RANK_VALUES = {"正量", "軽量", "過量"}


def _read_csv(file, **kwargs):
    for enc in ("cp932", "utf-8-sig"):
        try:
            return pd.read_csv(file, encoding=enc, **kwargs)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"CSVのエンコーディングを判別できませんでした: {os.path.basename(file)}")


def normalize_columns(file):

    df = _read_csv(file)
    df.columns = df.columns.str.replace("　", "").str.replace(" ", "").str.strip()
    df = df.loc[:, ~df.columns.duplicated()]
    cols = df.columns.tolist()

    # ===== アンリツ =====
    if any("測定値" in col for col in cols) and any("ランクコード" in col for col in cols):

        rename_map = {}
        for col in cols:
            if col == "測定値出力No.":
                rename_map[col] = "測定値出力No."
            elif col == "測定値(g)":
                rename_map[col] = "測定値(g)"
            elif "ランクコード" in col:
                rename_map[col] = "ランクコード"
            elif "日付時刻" in col:
                rename_map[col] = "日付時刻"

        df = df.rename(columns=rename_map)

        if "日付時刻" not in df.columns:
            if "日付" in df.columns and "時刻" in df.columns:
                df["日付時刻"] = pd.to_datetime(
                    df["日付"].astype(str) + " " + df["時刻"].astype(str),
                    errors="coerce"
                )

        hinshoku_num = None
        if "品種" in df.columns:
            vals = pd.to_numeric(df["品種"], errors="coerce").dropna()
            if len(vals) > 0:
                hinshoku_num = int(vals.iloc[0])

        df["メーカー"] = "アンリツ"
        df = df[["測定値出力No.", "日付時刻", "測定値(g)", "ランクコード", "メーカー"]].copy()
        df["測定値出力No."] = pd.to_numeric(df["測定値出力No."], errors="coerce")
        df["測定値(g)"] = pd.to_numeric(df["測定値(g)"], errors="coerce")
        df["ランクコード"] = df["ランクコード"].astype(str).str.strip()
        df["日付時刻"] = pd.to_datetime(df["日付時刻"], errors="coerce")
        return df, hinshoku_num

    # ===== イシダ判定 =====
    df_ishida = _read_csv(file, skiprows=10)
    df_ishida.columns = df_ishida.columns.str.replace("　", "").str.replace(" ", "").str.strip()

    is_ishida = (
        len(df_ishida.columns) >= 6
        and df_ishida.iloc[:, 5].dropna().astype(str).isin(_ISHIDA_RANK_VALUES).any()
    )
    if not is_ishida:
        raise ValueError(
            f"未対応のCSVフォーマットです。アンリツまたはイシダ形式のCSVを選択してください。\n"
            f"ファイル: {os.path.basename(file)}"
        )

    hinshoku_num = None
    if "予約番号" in df_ishida.columns:
        vals = pd.to_numeric(df_ishida["予約番号"], errors="coerce").dropna()
        if len(vals) > 0:
            hinshoku_num = int(vals.iloc[0])
    elif len(df_ishida.columns) >= 4:
        vals = pd.to_numeric(df_ishida.iloc[:, 3], errors="coerce").dropna()
        if len(vals) > 0:
            hinshoku_num = int(vals.iloc[0])

    df_ishida = df_ishida.iloc[:, [0, 1, 4, 5]].copy()
    df_ishida.columns = ["日付", "時刻", "測定値(g)", "判定"]

    df_ishida["日付時刻"] = pd.to_datetime(
        df_ishida["日付"].astype(str) + " " + df_ishida["時刻"].astype(str),
        errors="coerce"
    )
    df_ishida["測定値出力No."] = range(1, len(df_ishida) + 1)
    rank_map = {"正量": "2", "軽量": "1", "過量": "E"}
    df_ishida["ランクコード"] = df_ishida["判定"].map(rank_map)
    df_ishida["メーカー"] = "イシダ"

    return df_ishida[["測定値出力No.", "日付時刻", "測定値(g)", "ランクコード", "メーカー"]], hinshoku_num


# =========================
# ■ 分析
# =========================
def analyze(data):

    data = np.asarray(data).astype(float).ravel()
    data = data[~np.isnan(data)]

    n = len(data)
    if n < 2:
        return None, None, (None, None), None, None, None, None

    mean = float(np.mean(data))
    std = float(np.std(data, ddof=1))

    t_value = stats.t.ppf(0.975, df=n - 1)
    margin = t_value * std / np.sqrt(n)

    max1 = float(np.max(data))
    min1 = float(np.min(data))
    lower = mean - 3 * std
    upper = mean + 3 * std

    return mean, std, (mean - margin, mean + margin), max1, min1, lower, upper


# =========================
# ■ Excel出力
# =========================
def _create_report_sheet(wb, df_ok, mean, std, ci, max1, min1, lower, upper,
                          outliers_df, img_hist_bytes, img_series_bytes, rank_counts,
                          total_count, original_ok_count, hinshoku_num, date_str, lot,
                          img_all_series_bytes=None,
                          mfg_start="−", mfg_end="−", mfg_duration="−",
                          spec_nominal=None, usl=None, lsl=None,
                          lot_display=None, product_display=None):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image

    ws = wb.create_sheet("分析レポート", 0)

    # スタイル
    title_font  = Font(bold=True, size=16, color="FFFFFFFF")
    title_fill  = PatternFill(start_color="FF1F3864", fill_type="solid")
    sec_font    = Font(bold=True, size=10, color="FFFFFFFF")
    sec_fill    = PatternFill(start_color="FF4472C4", fill_type="solid")
    label_font  = Font(bold=True, size=10)
    label_fill  = PatternFill(start_color="FFD9E1F2", fill_type="solid")
    val_fill_e  = PatternFill(start_color="FFFFFFFF", fill_type="solid")
    val_fill_o  = PatternFill(start_color="FFEFF3FB", fill_type="solid")
    warn_fill   = PatternFill(start_color="FFFFF2CC", fill_type="solid")
    warn_font   = Font(bold=True, size=10, color="FF7F6000")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center", indent=1)

    # 列幅
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 12
    for col in ["E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col].width = 10.4  # E〜K合計 ≈ 13.5cm

    ng_count    = total_count - original_ok_count
    defect_rate = (ng_count / total_count * 100) if total_count > 0 else 0.0

    # タイトル行
    lot_display_str = lot_display if lot_display else f"ロット{lot}"
    if hinshoku_num is not None and date_str:
        prod_str = product_display or _hinshoku_display(hinshoku_num) or ""
        title = f"{date_str}製造   {prod_str}   {lot_display_str}   分析レポート"
    else:
        title = f"{lot_display_str}   分析レポート"

    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font      = title_font
    ws["A1"].fill      = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells("K1:L1")
    ws["K1"] = datetime.datetime.now().strftime("%Y年%m月%d日")
    ws["K1"].font      = Font(bold=True, size=11, color="FFFFFFFF")
    ws["K1"].fill      = title_fill
    ws["K1"].alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # 統計セクションヘッダー
    ws.merge_cells("A3:B3")
    ws["A3"] = "■ 基本統計"
    ws["A3"].font      = sec_font
    ws["A3"].fill      = sec_fill
    ws["A3"].alignment = center
    ws.row_dimensions[3].height = 16

    # 統計テーブル定義: (ラベル, 値, 書式, 警告フラグ)
    # ラベルが "__subheader__" の行は小見出しとして描画
    stats_rows = [
        ("全数",            total_count,       "0",     False),
        ("OK数",            original_ok_count, "0",     False),
        ("NG数",            ng_count,          "0",     ng_count > 0),
        ("不良率 (%)",       defect_rate,       "0.00",  ng_count > 0),
        ("__subheader__",  "製造情報",          None,    False),
        ("開始時刻",         mfg_start,         "@",     False),
        ("終了時刻",         mfg_end,           "@",     False),
        ("製造時間",         mfg_duration,      "@",     False),
        ("__subheader__",  "統計値（OK品のみ）", None,   False),
        ("平均 (g)",        mean,              "0.000", False),
    ]
    if spec_nominal is not None:
        deviation = mean - spec_nominal
        stats_rows += [
            ("基準値 (g)",        spec_nominal,      "0.000", False),
            ("基準値とのズレ (g)", deviation,         "+0.000", False),
        ]
    stats_rows += [
        ("標準偏差 (g)",     std,               "0.000", False),
        ("Max (g)",        max1,              "0.000", False),
        ("Min (g)",        min1,              "0.000", False),
        ("下限 −3σ (g)",   lower,             "0.000", False),
        ("上限 +3σ (g)",   upper,             "0.000", False),
        (None, None, None, False),
        ("外れ値件数",       len(outliers_df),  "0",     len(outliers_df) > 0),
    ]
    if usl is not None:
        stats_rows += [
            ("__subheader__", "規格値",     None,    False),
            ("UCL (g)",       usl,          "0.000", False),
            ("LCL (g)",       lsl,          "0.000", False),
        ]

    subheader_fill = PatternFill(start_color="FF8EA9C1", fill_type="solid")
    subheader_font = Font(bold=True, size=9, color="FFFFFFFF")

    data_row = 4
    stripe   = 0
    for label, value, fmt, warn in stats_rows:
        if label is None:
            ws.row_dimensions[data_row].height = 5
            data_row += 1
            continue
        if label == "__subheader__":
            ws.merge_cells(f"A{data_row}:B{data_row}")
            ws[f"A{data_row}"] = value
            ws[f"A{data_row}"].font      = subheader_font
            ws[f"A{data_row}"].fill      = subheader_fill
            ws[f"A{data_row}"].border    = thin
            ws[f"A{data_row}"].alignment = center
            ws.row_dimensions[data_row].height = 13
            data_row += 1
            continue

        ws.row_dimensions[data_row].height = 14
        vfill = warn_fill if warn else (val_fill_e if stripe % 2 == 0 else val_fill_o)
        vfont = warn_font if warn else Font(size=10)

        ws[f"A{data_row}"] = label
        ws[f"A{data_row}"].font      = label_font
        ws[f"A{data_row}"].fill      = label_fill
        ws[f"A{data_row}"].border    = thin
        ws[f"A{data_row}"].alignment = left

        ws[f"B{data_row}"] = value
        ws[f"B{data_row}"].font          = vfont
        ws[f"B{data_row}"].fill          = vfill
        ws[f"B{data_row}"].border        = thin
        ws[f"B{data_row}"].alignment     = center
        ws[f"B{data_row}"].number_format = fmt

        data_row += 1
        stripe   += 1

    ws.row_dimensions[data_row].height = 7  # 統計〜ランク間ギャップ（行3〜25計≈10.5cm調整）

    # ランクコード集計ミニテーブル
    rank_row = data_row + 1
    ws.merge_cells(f"A{rank_row}:C{rank_row}")
    ws[f"A{rank_row}"] = "■ ランクコード集計"
    ws[f"A{rank_row}"].font      = sec_font
    ws[f"A{rank_row}"].fill      = sec_fill
    ws[f"A{rank_row}"].alignment = center
    ws.row_dimensions[rank_row].height = 16

    rank_header_row = rank_row + 1
    for ci_idx, col_name in enumerate(["内容", "件数", "比率(%)"]):
        cell = ws.cell(row=rank_header_row, column=ci_idx + 1, value=col_name)
        cell.font      = Font(bold=True, color="FFFFFFFF")
        cell.fill      = PatternFill(start_color="FF4472C4", fill_type="solid")
        cell.border    = thin
        cell.alignment = center
    ws.row_dimensions[rank_header_row].height = 14

    for r_idx, row_data in rank_counts[["内容", "件数"]].iterrows():
        r = rank_header_row + 1 + r_idx
        rfill = val_fill_e if r_idx % 2 == 0 else val_fill_o
        ratio = row_data["件数"] / total_count * 100 if total_count > 0 else 0.0
        for c_idx, val in enumerate([row_data["内容"], row_data["件数"], ratio]):
            cell = ws.cell(row=r, column=c_idx + 1, value=val)
            cell.fill          = rfill
            cell.border        = thin
            cell.alignment     = center
            if c_idx == 2:
                cell.number_format = "0.00"
        ws.row_dimensions[r].height = 13

    # ヒストグラム (D3:J25 に TwoCellAnchor で固定 ≈ 13.5cm × 10.5cm)
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    img1 = Image(BytesIO(img_hist_bytes))
    img1.width  = 510  # fallback: 13.5cm
    img1.height = 397  # fallback: 10.5cm
    anchor1 = TwoCellAnchor(editAs="twoCell")
    anchor1._from = AnchorMarker(col=4, colOff=0, row=2,  rowOff=0)  # E3  (0-indexed)
    anchor1.to    = AnchorMarker(col=10, colOff=0, row=26, rowOff=0) # K27 (0-indexed)
    ws.add_image(img1, anchor1)

    # 時系列チャート (下段) ヒストグラムアンカー to=row26(0-indexed) → Excel行27まで占有。行28以降に配置する
    first_series_row = max(data_row, rank_header_row + len(rank_counts) + 1, 26) + 2

    if img_all_series_bytes is not None:
        # ① 全データ時系列
        ws.merge_cells(f"A{first_series_row}:L{first_series_row}")
        ws[f"A{first_series_row}"] = "■ 全データ時系列"
        ws[f"A{first_series_row}"].font      = sec_font
        ws[f"A{first_series_row}"].fill      = sec_fill
        ws[f"A{first_series_row}"].alignment = center
        ws.row_dimensions[first_series_row].height = 16

        img3 = Image(BytesIO(img_all_series_bytes))
        img3.width  = 907
        img3.height = 378
        ws.add_image(img3, f"A{first_series_row + 1}")

        # ② 時系列チャート（OK品のみ）
        series_row = first_series_row + 26
    else:
        series_row = first_series_row

    ws.merge_cells(f"A{series_row}:L{series_row}")
    ws[f"A{series_row}"] = "■ 時系列チャート（OK品のみ）"
    ws[f"A{series_row}"].font      = sec_font
    ws[f"A{series_row}"].fill      = sec_fill
    ws[f"A{series_row}"].alignment = center
    ws.row_dimensions[series_row].height = 16

    img2 = Image(BytesIO(img_series_bytes))
    img2.width  = 907
    img2.height = 378
    ws.add_image(img2, f"A{series_row + 1}")

    # A4縦・幅1ページ印刷設定（高さは自動）
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.6
    ws.page_margins.right = 0.6
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3


def save_to_excel(df_ok, mean, std, ci, max1, min1, lower, upper,
                  outliers_df, img_hist, img_series, rank_counts, filename, lot,
                  total_count=0, original_ok_count=0, hinshoku_num=None, date_str=None,
                  df_all=None, img_all_series=None,
                  mfg_start="−", mfg_end="−", mfg_duration="−",
                  spec_nominal=None, usl=None, lsl=None,
                  lot_display=None, product_display=None):

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image

    red_fill    = PatternFill(start_color="FFFF0000", fill_type="solid")
    header_fill = PatternFill(start_color="FF4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    rank_label_map = {"2": "OK", "1": "軽量", "E": "過量", "0": "２個乗り"}

    df_all_out = (df_all if df_all is not None else df_ok)[
        ["測定値出力No.", "日付時刻", "測定値(g)", "ランクコード"]
    ].copy()
    df_all_out["ランクコード"] = df_all_out["ランクコード"].map(rank_label_map).fillna(df_all_out["ランクコード"])

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        df_ok[["測定値出力No.", "日付時刻", "測定値(g)"]].to_excel(writer, sheet_name="OKデータ", index=False)
        outliers_df[["測定値出力No.", "日付時刻", "測定値(g)"]].to_excel(writer, sheet_name="外れ値（±3σ超）", index=False)
        df_all_out.to_excel(writer, sheet_name="全データ", index=False)

        wb = writer.book

        # ===== OKデータシート =====
        ws_ok = wb["OKデータ"]
        ws_ok.column_dimensions["A"].width = 18
        ws_ok.column_dimensions["B"].width = 22
        ws_ok.column_dimensions["C"].width = 15
        for cell in ws_ok[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        for row in ws_ok.iter_rows(min_row=2, min_col=1, max_col=3):
            for cell in row:
                cell.border = border
        for row in ws_ok.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        ws_ok.auto_filter.ref = f"A1:C{ws_ok.max_row}"
        outlier_ids = set(outliers_df["測定値出力No."].values)
        for row in ws_ok.iter_rows(min_row=2, max_row=ws_ok.max_row):
            if row[0].value in outlier_ids:
                for cell in row:
                    cell.fill = red_fill

        # ===== 外れ値シート =====
        ws_out = wb["外れ値（±3σ超）"]
        ws_out.column_dimensions["A"].width = 18
        ws_out.column_dimensions["B"].width = 22
        ws_out.column_dimensions["C"].width = 15
        for cell in ws_out[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        for row in ws_out.iter_rows(min_row=2, min_col=1, max_col=3):
            for cell in row:
                cell.border = border
        for row in ws_out.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        ws_out.auto_filter.ref = f"A1:C{ws_out.max_row}"

        # ===== 全データシート =====
        ws_all = wb["全データ"]
        ws_all.column_dimensions["A"].width = 18
        ws_all.column_dimensions["B"].width = 22
        ws_all.column_dimensions["C"].width = 15
        ws_all.column_dimensions["D"].width = 12
        for cell in ws_all[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        ng_labels = {"軽量", "過量"}
        for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row):
            is_ng = row[3].value in ng_labels  # D列 = ランクコード
            for cell in row:
                cell.border = border
                cell.alignment = center_align
                if is_ng:
                    cell.fill = red_fill
            row[1].number_format = "yyyy-mm-dd hh:mm:ss"  # B列 = 日付時刻
        ws_all.auto_filter.ref = f"A1:D{ws_all.max_row}"

        # ===== グラフシート =====
        img_hist_bytes   = img_hist.getvalue()
        img_series_bytes = img_series.getvalue()

        ws_hist = wb.create_sheet("ヒストグラム（OK品のみ）")
        ws_hist.add_image(Image(BytesIO(img_hist_bytes)), "A1")

        ws_series = wb.create_sheet("時系列チャート（OK品のみ）")
        ws_series.add_image(Image(BytesIO(img_series_bytes)), "A1")

        if img_all_series is not None:
            ws_all_series = wb.create_sheet("全データ時系列")
            ws_all_series.add_image(Image(BytesIO(img_all_series.getvalue())), "A1")

        # ===== 分析レポートシート =====
        _create_report_sheet(
            wb, df_ok, mean, std, ci, max1, min1, lower, upper,
            outliers_df, img_hist_bytes, img_series_bytes, rank_counts,
            total_count, original_ok_count, hinshoku_num, date_str, lot,
            img_all_series_bytes=img_all_series.getvalue() if img_all_series is not None else None,
            mfg_start=mfg_start, mfg_end=mfg_end, mfg_duration=mfg_duration,
            spec_nominal=spec_nominal, usl=usl, lsl=lsl,
            lot_display=lot_display, product_display=product_display
        )

        # ===== シート順序を整理 =====
        sheet_order = [
            "分析レポート",
            "全データ時系列",
            "全データ",
            "ヒストグラム（OK品のみ）",
            "時系列チャート（OK品のみ）",
            "OKデータ",
            "外れ値（±3σ超）",
        ]
        wb._sheets.sort(
            key=lambda ws: sheet_order.index(ws.title) if ws.title in sheet_order else len(sheet_order)
        )


# =========================
# ■ ロット処理
# =========================
def process_lot(group, lot, save_dir, hinshoku_num=None, spec=None, lot_label=None,
                product_name=None, ma_window=MA_WINDOW):
    """
    1ロット分の分析を行いExcelを出力する。
    spec: None | (nominal_or_None, offset_or_None)
    戻り値:
        ("ok",   ok_count) … 正常に作成
        ("skip", ok_count) … OKデータ不足によりスキップ
    """

    rank_map = {"2": "OK", "1": "軽量", "E": "過量", "0": "２個乗り"}

    rank_counts = group["ランクコード"].value_counts().reset_index()
    rank_counts.columns = ["ランクコード", "件数"]
    rank_counts["内容"] = rank_counts["ランクコード"].map(rank_map)
    rank_counts = rank_counts[["ランクコード", "内容", "件数"]]

    total_count = len(group)
    original_ok_count = int((group["ランクコード"] == "2").sum())

    df_ok = group[group["ランクコード"] == "2"].copy()

    data = pd.to_numeric(df_ok["測定値(g)"], errors="coerce")
    df_ok = df_ok.loc[data.notna()].copy()
    data = data.loc[data.notna()]
    data = np.asarray(data).ravel()

    if len(data) < MIN_OK_COUNT:
        return ("skip", len(data))

    mean, std, ci, max1, min1, lower, upper = analyze(data)

    outliers_df = df_ok[(df_ok["測定値(g)"] < lower) | (df_ok["測定値(g)"] > upper)]

    # ===== 規格値・工程能力 =====
    spec_nominal = usl = lsl = None
    if spec is not None:
        input_nominal, upper_offset, lower_offset = spec
        # 基準値：入力値 or 先頭5個OK品の平均
        if input_nominal is not None:
            spec_nominal = input_nominal
        else:
            first5 = df_ok.sort_values("日付時刻").head(5)
            spec_nominal = pd.to_numeric(first5["測定値(g)"], errors="coerce").mean()
        # UCL = 基準値 + 上側許容幅、LCL = 基準値 − 下側許容幅（非対称対応）
        usl = spec_nominal + upper_offset if upper_offset is not None else None
        lsl = spec_nominal - lower_offset if lower_offset is not None else None

    lot_display = lot_label if lot_label else f"ロット{lot}"
    lot_display_safe = lot_display.translate(str.maketrans('\\/:*?"<>|', '_________'))

    lot_date = group["日付時刻"].dropna().min()
    lot_end  = group["日付時刻"].dropna().max()
    if pd.notna(lot_date) and pd.notna(lot_end):
        mfg_start    = lot_date.strftime("%H:%M")
        mfg_end      = lot_end.strftime("%H:%M")
        total_min    = int((lot_end - lot_date).total_seconds() / 60)
        mfg_duration = f"{total_min // 60}時間{total_min % 60:02d}分"
    else:
        mfg_start = mfg_end = mfg_duration = "−"

    if pd.notna(lot_date) and hinshoku_num is not None:
        date_str = f"{lot_date.year}/{lot_date.month}/{lot_date.day}"
        product_display = product_name if product_name else _hinshoku_display(hinshoku_num)
        chart_prefix = f"{date_str}製造 {product_display} "
        date_str_safe = date_str.replace("/", "-")
    else:
        date_str = None
        product_display = product_name or None
        chart_prefix = f"{product_display} " if product_display else ""
        date_str_safe = None

    # ===== ヒストグラム =====
    fig1, ax1 = plt.subplots(figsize=(10, 6))
    ax1.hist(data, bins=30, edgecolor="black", alpha=0.7)
    ax1.axvline(mean, color="red", linestyle="-", linewidth=2, label=f"平均: {mean:.3f}")
    if spec_nominal is not None:
        ax1.axvline(spec_nominal, color="navy", linewidth=2.0, linestyle=":",
                    label=f"基準値: {spec_nominal:.3f}")
    ax1.axvline(lower, color="orange", linestyle="--", linewidth=2, label=f"下限(-3σ): {lower:.3f}")
    ax1.axvline(upper, color="orange", linestyle="--", linewidth=2, label=f"上限(+3σ): {upper:.3f}")
    if usl is not None:
        ax1.axvline(usl, color="purple", linewidth=2.0, linestyle="-.", label=f"UCL: {usl:.3f}")
        ax1.axvline(lsl, color="purple", linewidth=2.0, linestyle="-.", label=f"LCL: {lsl:.3f}")
    ax1.set_title(f"{chart_prefix}{lot_display}　測定値の分布（OK品のみ・n={len(data)}）", fontsize=14, fontweight="bold")
    ax1.set_xlabel("測定値(g)", fontsize=12)
    ax1.set_ylabel("頻度", fontsize=12)
    ax1.legend(fontsize=10)
    ax1.grid(True, alpha=0.3)
    fig1.tight_layout()

    img_hist = BytesIO()
    fig1.savefig(img_hist, format="png", dpi=100, bbox_inches="tight")
    img_hist.seek(0)
    plt.close(fig1)

    # ===== 時系列チャート =====
    fig2, ax2 = plt.subplots(figsize=(12, 5))

    y_vals = df_ok["測定値(g)"].values
    outlier_mask = (df_ok["測定値(g)"] < lower) | (df_ok["測定値(g)"] > upper)
    has_datetime = df_ok["日付時刻"].notna().any()

    if has_datetime:
        x_all = df_ok["日付時刻"]
        x_ok  = df_ok.loc[~outlier_mask, "日付時刻"]
        x_out = df_ok.loc[outlier_mask,  "日付時刻"]
        ax2.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        ax2.xaxis.set_major_locator(mdates.AutoDateLocator())
        plt.setp(ax2.xaxis.get_majorticklabels(), rotation=30, ha="right")
        ax2.set_xlabel("時刻", fontsize=12)
    else:
        x_all = np.arange(1, len(df_ok) + 1)
        x_ok  = x_all[~outlier_mask.values]
        x_out = x_all[outlier_mask.values]
        ax2.set_xlabel("測定順序", fontsize=12)

    y_ok  = y_vals[~outlier_mask.values]
    y_out = y_vals[outlier_mask.values]

    ax2.plot(x_all, y_vals, color="steelblue", linewidth=0.6, alpha=0.4, zorder=1)
    ax2.scatter(x_ok, y_ok, color="steelblue", s=18, alpha=0.8, zorder=2, label="OK")
    if len(x_out) > 0:
        ax2.scatter(x_out, y_out, color="red", s=50, marker="x",
                    linewidths=2, zorder=3, label=f"外れ値・±3σ超 ({len(x_out)}件)")

    ax2.axhline(mean,  color="red",    linewidth=1.5, linestyle="-",  label=f"平均: {mean:.3f}")
    if spec_nominal is not None:
        ax2.axhline(spec_nominal, color="navy", linewidth=1.5, linestyle=":",
                    label=f"基準値: {spec_nominal:.3f}")
    ax2.axhline(upper, color="orange", linewidth=1.5, linestyle="--", label=f"+3σ: {upper:.3f}")
    ax2.axhline(lower, color="orange", linewidth=1.5, linestyle="--", label=f"-3σ: {lower:.3f}")
    if usl is not None:
        ax2.axhline(usl, color="purple", linewidth=2.0, linestyle="-.", label=f"UCL: {usl:.3f}")
        ax2.axhline(lsl, color="purple", linewidth=2.0, linestyle="-.", label=f"LCL: {lsl:.3f}")
    ma2 = pd.Series(y_vals).rolling(window=ma_window, min_periods=1).mean()
    ax2.plot(x_all, ma2.values, color="green", linewidth=2.0,
             alpha=0.9, zorder=5, label=f"移動平均（{ma_window}点）")

    ax2.set_title(f"{chart_prefix}{lot_display}　時系列チャート（OK品のみ・n={len(df_ok)}）", fontsize=14, fontweight="bold")
    ax2.set_ylabel("測定値(g)", fontsize=12)
    ax2.legend(fontsize=10)
    ax2.grid(True, alpha=0.3)
    fig2.tight_layout()

    img_series = BytesIO()
    fig2.savefig(img_series, format="png", dpi=100, bbox_inches="tight")
    img_series.seek(0)
    plt.close(fig2)

    # ===== 全データ時系列チャート =====
    group_sorted = group.sort_values("日付時刻").reset_index(drop=True)
    ok_mask    = group_sorted["ランクコード"] == "2"
    kacho_mask = group_sorted["ランクコード"] == "E"
    keiry_mask = group_sorted["ランクコード"] == "1"

    y_vals_all  = pd.to_numeric(group_sorted["測定値(g)"], errors="coerce")
    y_ok_all    = y_vals_all[ok_mask].values
    y_kacho_all = y_vals_all[kacho_mask].values
    y_keiry_all = y_vals_all[keiry_mask].values

    fig3, ax3 = plt.subplots(figsize=(12, 5))

    if group_sorted["日付時刻"].notna().any():
        x_ok_all    = group_sorted.loc[ok_mask,    "日付時刻"]
        x_kacho_all = group_sorted.loc[kacho_mask, "日付時刻"]
        x_keiry_all = group_sorted.loc[keiry_mask, "日付時刻"]
        x_all_for_ma = group_sorted["日付時刻"]
        ax3.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        ax3.xaxis.set_major_locator(mdates.AutoDateLocator())
        plt.setp(ax3.xaxis.get_majorticklabels(), rotation=30, ha="right")
        ax3.set_xlabel("時刻", fontsize=12)
    else:
        x_base       = np.arange(1, len(group_sorted) + 1)
        x_ok_all     = x_base[ok_mask.values]
        x_kacho_all  = x_base[kacho_mask.values]
        x_keiry_all  = x_base[keiry_mask.values]
        x_all_for_ma = x_base
        ax3.set_xlabel("測定順序", fontsize=12)

    ax3.plot(x_ok_all, y_ok_all, color="steelblue", linewidth=0.6, alpha=0.4, zorder=1)
    ax3.scatter(x_ok_all, y_ok_all, color="steelblue", s=18, alpha=0.8, zorder=2,
                label=f"OK ({ok_mask.sum()}件)")
    if kacho_mask.any():
        ax3.scatter(x_kacho_all, y_kacho_all, color="red", s=60, marker="^", zorder=4,
                    label=f"過量 ({kacho_mask.sum()}件)")
    if keiry_mask.any():
        ax3.scatter(x_keiry_all, y_keiry_all, color="orange", s=60, marker="v", zorder=4,
                    label=f"軽量 ({keiry_mask.sum()}件)")

    ax3.axhline(mean,  color="red",    linewidth=1.5, linestyle="-",  label=f"平均（OK品）: {mean:.3f}")
    if spec_nominal is not None:
        ax3.axhline(spec_nominal, color="navy", linewidth=1.5, linestyle=":",
                    label=f"基準値: {spec_nominal:.3f}")
    ax3.axhline(upper, color="darkorange", linewidth=1.5, linestyle="--", label=f"+3σ（OK品）: {upper:.3f}")
    ax3.axhline(lower, color="darkorange", linewidth=1.5, linestyle="--", label=f"-3σ（OK品）: {lower:.3f}")
    ma3_all = y_vals_all.rolling(window=ma_window, min_periods=1).mean()
    ax3.plot(x_all_for_ma, ma3_all.values, color="green", linewidth=2.0,
             alpha=0.9, zorder=5, label=f"移動平均・全数（{ma_window}点）")
    if usl is not None:
        ax3.axhline(usl, color="purple", linewidth=2.0, linestyle="-.", label=f"UCL: {usl:.3f}")
        ax3.axhline(lsl, color="purple", linewidth=2.0, linestyle="-.", label=f"LCL: {lsl:.3f}")
    y_lo = min(mean - 5 * std, lsl - std * 0.5) if lsl is not None else mean - 5 * std
    y_hi = max(mean + 5 * std, usl + std * 0.5) if usl is not None else mean + 5 * std
    ax3.set_ylim(y_lo, y_hi)

    ax3.set_title(f"{chart_prefix}{lot_display}　全データ時系列（n={len(group_sorted)}）", fontsize=14, fontweight="bold")
    ax3.set_ylabel("測定値(g)", fontsize=12)
    ax3.legend(fontsize=10)
    ax3.grid(True, alpha=0.3)
    fig3.tight_layout()

    img_all_series = BytesIO()
    fig3.savefig(img_all_series, format="png", dpi=100, bbox_inches="tight")
    img_all_series.seek(0)
    plt.close(fig3)

    if date_str_safe and hinshoku_num is not None:
        filename = os.path.join(
            save_dir,
            f"分析結果_{date_str_safe}製造_{product_display} {lot_display_safe}_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx"
        )
    else:
        filename = os.path.join(
            save_dir,
            f"分析結果_{lot_display_safe}_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx"
        )

    save_to_excel(df_ok, mean, std, ci, max1, min1,
                  lower, upper, outliers_df,
                  img_hist, img_series, rank_counts, filename, lot,
                  total_count=total_count, original_ok_count=original_ok_count,
                  hinshoku_num=hinshoku_num, date_str=date_str,
                  df_all=group, img_all_series=img_all_series,
                  mfg_start=mfg_start, mfg_end=mfg_end, mfg_duration=mfg_duration,
                  spec_nominal=spec_nominal, usl=usl, lsl=lsl,
                  lot_display=lot_display, product_display=product_display)

    return ("ok", len(data))


# =========================
# ■ ファイル処理
# =========================
def process_files(files):
    try:
        if not files:
            return

        for f in files:
            if not os.path.exists(f):
                messagebox.showerror("エラー", f"ファイルが見つかりません: {os.path.basename(f)}")
                return

        save_dir = os.path.dirname(files[0])

        df_list = []
        hinshoku_num = None
        for f in files:
            df_part, hnum = normalize_columns(f)
            df_list.append(df_part)
            if hinshoku_num is None:
                hinshoku_num = hnum
            elif hnum is not None and hnum != hinshoku_num:
                if not messagebox.askyesno(
                    "品種番号不一致",
                    f"品種番号が異なるファイルが含まれています。\n"
                    f"（{hinshoku_num} と {hnum}）\n\n続けてよいですか？"
                ):
                    return

        df = pd.concat(df_list, ignore_index=True).sort_values("日付時刻").reset_index(drop=True)
        df["測定値出力No."] = range(1, len(df) + 1)

        dialog = LotPreviewDialog(app_root, df, hinshoku_num)
        app_root.wait_window(dialog)

        if dialog.result is None or dialog.result[0] == "cancel":
            return
        if dialog.result[0] == "manual":
            messagebox.showwarning("中止", "CSVを手動分割してください")
            return

        df = dialog.result[1]
        ma_window = getattr(dialog, "ma_window", MA_WINDOW)

        # ロットごとに規格値入力
        spec_per_lot = {}      # lot -> (spec, lot_label, product_name)
        for lot in sorted(df["ロット"].unique()):
            g = df[df["ロット"] == lot]
            ok_count_lot = int((g["ランクコード"] == "2").sum())
            if ok_count_lot < MIN_OK_COUNT:
                spec_per_lot[lot] = (None, None, None)
                continue
            spec_dialog = SpecInputDialog(app_root, hinshoku_num,
                                          lot=lot, ok_count=ok_count_lot)
            app_root.wait_window(spec_dialog)
            if spec_dialog.result == "skip":
                spec_per_lot[lot] = (None, None, None)
            else:
                spec_per_lot[lot] = (spec_dialog.result, spec_dialog.lot_label,
                                     spec_dialog.product_name)

        # 結果集計
        created_lots = []   # [(lot, ok_count), ...]
        skipped_lots = []   # [(lot, ok_count, total_count), ...]

        for lot, group in df.groupby("ロット"):
            total = len(group)
            lot_spec, lot_label, product_name = spec_per_lot.get(lot, (None, None, None))
            status, ok_count = process_lot(group, lot, save_dir, hinshoku_num,
                                           spec=lot_spec, lot_label=lot_label,
                                           product_name=product_name,
                                           ma_window=ma_window)
            if status == "ok":
                created_lots.append((lot, ok_count))
            else:
                skipped_lots.append((lot, ok_count, total))

        # ===== 完了メッセージ =====
        if not created_lots and skipped_lots:
            msg = "Excelファイルは作成されませんでした。\n\n"
            msg += "すべてのロットでOKデータが不足しています:\n"
            for lot, ok, total in skipped_lots:
                msg += f"  ・ロット{lot}: 総{total}件 / OK{ok}件\n"
            msg += "\nしきい値を変更するか、CSVの内容をご確認ください。"
            messagebox.showerror("作成失敗", msg)

        elif skipped_lots:
            msg = f"Excel作成完了\n作成: {len(created_lots)}ファイル\n\n"
            msg += "⚠ 以下のロットはOKデータ不足のためスキップしました:\n"
            for lot, ok, total in skipped_lots:
                msg += f"  ・ロット{lot}: 総{total}件 / OK{ok}件\n"
            msg += f"\n（OKデータが {MIN_OK_COUNT} 件未満のロットは統計計算ができません）"
            messagebox.showwarning("完了（一部スキップ）", msg)

        else:
            messagebox.showinfo(
                "完了",
                f"Excel作成完了\n{len(created_lots)}ファイルを作成しました。"
            )

    except Exception as e:
        messagebox.showerror("エラー", str(e))


def process_file(file):
    process_files([file])


# =========================
# ■ メイン
# =========================
def run():
    files = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
    if files:
        process_files(list(files))


def on_closing():
    if app_root:
        app_root.destroy()


if __name__ == "__main__":
    app_root = tk.Tk()
    app_root.title("重量分析ツール")
    app_root.protocol("WM_DELETE_WINDOW", on_closing)

    btn = tk.Button(app_root, text="CSV選択して解析", command=run, height=2, width=30)
    btn.pack(pady=20)

    app_root.update_idletasks()
    w = app_root.winfo_width()
    h = app_root.winfo_height()
    sw = app_root.winfo_screenwidth()
    sh = app_root.winfo_screenheight()
    app_root.geometry(f"+{(sw - w) // 2}+{(sh - h) // 2}")

    if len(sys.argv) > 1:
        csv_file_to_process = sys.argv[1]
        app_root.after(500, lambda: process_file(csv_file_to_process))

    app_root.mainloop()
